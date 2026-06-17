import streamlit as st
import pandas as pd
import numpy as np
import sqlite3
import os
import re
from datetime import datetime, date, time, timedelta
import random

import plotly.graph_objects as go

st.set_page_config(page_title="Proline Analytics — Demo", layout="wide", page_icon="📊")

st.markdown("""
<style>
    header[data-testid="stHeader"] {display: none;}
    footer {visibility: hidden;} #MainMenu {visibility: hidden;}
    html, body, [class*="css"] {font-family: 'Inter','Segoe UI',-apple-system,BlinkMacSystemFont,sans-serif;}
    .proline-header {background:#FFFFFF;padding:16px 28px;margin:-6rem -4rem 1.5rem -4rem;border-bottom:1px solid #E8ECF0;display:flex;align-items:center;gap:16px;}
    .proline-header .brand-text h1 {color:#1B3A5C;font-size:22px;font-weight:700;margin:0;letter-spacing:-0.3px;}
    .proline-header .brand-text .subtitle {color:#778899;font-size:12px;font-weight:400;margin-top:1px;}
    .kpi-card {background:#FFFFFF;border-radius:12px;padding:18px 16px 14px 16px;margin:4px 2px;box-shadow:0 2px 8px rgba(0,0,0,0.06);border-top:4px solid #E0E0E0;position:relative;}
    .kpi-card.bad {border-top-color:#C62828;background:#FFF5F5;}
    .kpi-card.good {border-top-color:#2E7D32;background:#F5FDF5;}
    .kpi-card.warn {border-top-color:#E65100;background:#FFF8F0;}
    .kpi-card.neutral {border-top-color:#2196F3;background:#F8FBFF;}
    .kpi-label {font-size:11px;text-transform:uppercase;letter-spacing:1px;color:#666;margin-bottom:6px;}
    .kpi-value {font-size:30px;font-weight:700;color:#1B3A5C;line-height:1.1;}
    .kpi-unit {font-size:12px;color:#999;font-weight:400;margin-left:2px;}
    .kpi-delta {font-size:12px;margin-top:4px;}
    .status-dot {display:inline-block;width:8px;height:8px;border-radius:50%;margin-right:6px;}
    .status-dot.red {background:#C62828;} .status-dot.green {background:#2E7D32;}
    .status-dot.orange {background:#E65100;} .status-dot.blue {background:#2196F3;}
    .comparison-current {background:#E3F2FD;border-left:4px solid #2196F3;padding:8px 12px;border-radius:6px;}
    .comparison-prev {background:#F5F5F5;border-left:4px solid #9E9E9E;padding:8px 12px;border-radius:6px;}
    .comp-hero {border-radius:16px;padding:24px 28px;margin:12px 0 20px 0;box-shadow:0 4px 16px rgba(0,0,0,0.08);text-align:center;}
    .comp-hero.worse {background:linear-gradient(135deg,#FFF5F5,#FFEBEE);border:1px solid #FFCDD2;}
    .comp-hero.better {background:linear-gradient(135deg,#F1F8E9,#E8F5E9);border:1px solid #C8E6C9;}
    .comp-hero .hero-label {font-size:13px;text-transform:uppercase;letter-spacing:1.5px;margin-bottom:6px;}
    .comp-hero .hero-value {font-size:42px;font-weight:800;line-height:1.1;}
    .comp-hero.worse .hero-label {color:#B71C1C;} .comp-hero.worse .hero-value {color:#C62828;}
    .comp-hero.better .hero-label {color:#1B5E20;} .comp-hero.better .hero-value {color:#2E7D32;}
    .comp-hero .hero-detail {font-size:13px;color:#666;margin-top:6px;}
    .stat-row {display:flex;gap:12px;margin:12px 0;}
    .stat-row .stat-card {flex:1;background:#FFFFFF;border-radius:12px;padding:16px 18px;box-shadow:0 2px 8px rgba(0,0,0,0.05);border:1px solid #E8ECF0;}
    .stat-row .stat-card .stat-label {font-size:11px;text-transform:uppercase;letter-spacing:1px;color:#888;}
    .stat-row .stat-card .stat-num {font-size:26px;font-weight:700;color:#1B3A5C;}
    .stat-row .stat-card .stat-sub {font-size:12px;color:#999;margin-top:2px;}
</style>
""", unsafe_allow_html=True)

REJECT_CODES = {2, 3, 4, 5, 6, 7}

@st.cache_resource
def create_demo_database():
    random.seed(42); np.random.seed(42)
    MACHINES = ["IIN2-053","PNP4-101","YKM3-087"]
    BOARDS = ["MAIN-BOARD-A","MAIN-BOARD-B","PWR-MODULE-C","CTRL-PANEL-D"]
    MO_NUMBERS = ["MO-2026-001","MO-2026-002","MO-2026-003"]
    FEEDERS = [f"F{i}" for i in range(1, 16)]
    SLOTS = [f"S{i}" for i in range(1, 11)]
    COMPONENTS = {
        "RES-10K":("Panasonic","Resistor 10K 5% 0603",0.01),"RES-22K":("Panasonic","Resistor 22K 1% 0603",0.01),
        "RES-47K":("Panasonic","Resistor 47K 5% 0603",0.02),"RES-100K":("Yageo","Resistor 100K 5% 0805",0.02),
        "RES-1M":("Yageo","Resistor 1M 5% 0805",0.03),"CAP-10UF":("Murata","Capacitor 10uF X5R 25V",0.05),
        "CAP-22UF":("Murata","Capacitor 22uF X7R 16V",0.06),"CAP-47UF":("TDK","Capacitor 47uF X7R 10V",0.08),
        "CAP-100UF":("AVX","Capacitor 100uF X5R 6.3V",0.12),"IC-555":("TI","IC Timer NE555P DIP-8",0.45),
        "IC-741":("TI","IC Op-Amp LM741CN DIP-8",0.38),"IC-7805":("TI","IC Regulator 7805 TO-220",0.52),
        "IC-AT328":("Microchip","IC MCU ATmega328P TQFP-32",1.85),"DIODE-1N4":("Vishay","Diode 1N4007 Rectifier",0.03),
        "DIODE-ZEN1":("Vishay","Diode Zener 1N4733A 5.1V",0.08),"DIODE-SCH":("ON Semi","Diode Schottky 1N5819",0.11),
        "CONN-HEAD8":("Samtec","Connector Header 8P 2.54mm",0.25),"CONN-HEAD4":("Samtec","Connector Header 4P 2.54mm",0.18),
        "CONN-FPC20":("Molex","Connector FPC 20P 0.5mm",0.65),"XTAL-16MHZ":("Abracon","Crystal 16MHz SMD",0.35),
        "XTAL-8MHZ":("Abracon","Crystal 8MHz SMD",0.30),"LED-RED":("Kingbright","LED Red 0805",0.04),
        "LED-GRN":("Kingbright","LED Green 0805",0.04),"FUSE-1A":("Littelfuse","Fuse 1A SMD 1206",0.15),
        "FUSE-2A":("Littelfuse","Fuse 2A SMD 1206",0.18),
        "MCU-STM32":("STMicro","MCU STM32F407VGT6 LQFP-100",8.50),
        "FPGA-SPART":("Xilinx","FPGA Spartan-6 XC6SLX9 TQG-144",9.75),
        "DCDC-5V":("Murata","DC-DC Converter 5V 3A Module",6.20),
        "RF-TRANS":("Nordic","RF Transceiver nRF24L01+ QFN-20",5.50),
        "ADC-12BIT":("TI","ADC 12-Bit ADS1015 QFN",7.80),
    }
    COMPONENT_NAMES = list(COMPONENTS.keys())
    # Assign spit weights — expensive parts get high weights to generate large rejection costs
    COMPONENT_SPIT_WEIGHTS = {}
    expensive_high_weight = {"MCU-STM32", "FPGA-SPART", "DCDC-5V", "RF-TRANS", "ADC-12BIT",
                             "IC-AT328", "CONN-FPC20", "IC-7805", "IC-741", "IC-555"}
    cheap_wts = [50,40,35,30,25,22,20,18,15,12,10,8,7,6,5,4,3,3,2,2]
    random.shuffle(cheap_wts)
    ci = 0
    for comp in COMPONENT_NAMES:
        if comp in expensive_high_weight:
            COMPONENT_SPIT_WEIGHTS[comp] = random.choice([200, 250, 300, 350, 400, 450, 500, 550, 600, 700])
        else:
            COMPONENT_SPIT_WEIGHTS[comp] = cheap_wts[ci] if ci < len(cheap_wts) else 1
            ci += 1
    tw = sum(COMPONENT_SPIT_WEIGHTS.get(c,1) for c in COMPONENT_NAMES)
    comp_probs_norm = [COMPONENT_SPIT_WEIGHTS.get(c,1)/tw for c in COMPONENT_NAMES]
    COMPONENT_REJECT_CODES = {}
    for comp in COMPONENT_NAMES: COMPONENT_REJECT_CODES[comp] = random.sample(sorted(REJECT_CODES), random.randint(1,3))
    MACHINE_SPIT_RATE = {"IIN2-053":(6,12),"PNP4-101":(4,9),"YKM3-087":(9,18)}
    logs, events = [], []
    today = date(2026,6,10); start_date = today - timedelta(days=30)
    file_hash_counter = 0
    for machine in MACHINES:
        for day_offset in range(30):
            dt_val = start_date + timedelta(days=day_offset)
            num_logs = random.randint(50,70)
            for _ in range(num_logs):
                file_hash_counter += 1
                file_hash = f"demo_hash_{file_hash_counter:07d}"
                board_name = random.choice(BOARDS); mo = random.choice(MO_NUMBERS)
                h = random.randint(0,23); m = random.randint(0,59); s2 = random.randint(0,59)
                dt_iso = f"{dt_val.isoformat()} {h:02d}:{m:02d}:{s2:02d}"
                filename = f"{dt_val.strftime('%Y%m%d')}{h:02d}{m:02d}{s2:02d}-{machine}-{file_hash_counter}.csv"
                logs.append({"file_hash":file_hash,"filename":filename,"file_dt":dt_iso,"machine":machine,"board_name":board_name,"mo":mo,"source_path":f"C:\\DemoLogs\\{machine}\\{filename}","source_mtime":0.0,"source_size":0,"ingested_at":dt_iso})
                nc = random.randint(120,180); lr,hr = MACHINE_SPIT_RATE[machine]
                spit_pct = random.randint(lr,hr)
                n_spits = max(0, int(nc*spit_pct/100.0)); n_success = nc - n_spits
                spit_comps = random.choices(COMPONENT_NAMES, weights=comp_probs_norm, k=n_spits)
                success_comps = random.choices(COMPONENT_NAMES, k=n_success)
                sampled = spit_comps + success_comps; random.shuffle(sampled); sampled = sampled[:nc]
                for idx, comp in enumerate(sampled):
                    man, desc_txt, uc = COMPONENTS[comp]
                    is_spit = comp in set(spit_comps[:len(sampled)])
                    rc = random.choice(COMPONENT_REJECT_CODES[comp]) if is_spit else 0
                    feeder_no = random.choice(FEEDERS); slot_no = random.choice(SLOTS)
                    location = f"LOC-{random.randint(100,999)}"
                    events.append({"file_hash":file_hash,"source_row_index":idx,"component":comp.upper(),"description":f"{man} {desc_txt}","location":location,"feeder_no":feeder_no,"slot_no":slot_no,"board_name":board_name,"mo":mo,"file_dt":dt_iso,"machine":machine,"unit_cost":uc,"cost":uc,"reject_code":rc})
    # Cap PER DAY per component — ensures every day shows meaningful totals (£40-£150)
    cap_pool = [40, 60, 60, 60, 60, 60, 70, 70, 70, 70, 80, 80, 100, 100, 120, 120, 140, 140, 150, 150, 150, 150, 150, 150, 150, 150, 150, 150, 150, 150]
    random.shuffle(cap_pool)
    COMPONENT_SPIT_CAPS = {}
    for i, comp in enumerate(COMPONENT_NAMES):
        COMPONENT_SPIT_CAPS[comp] = cap_pool[i] if i < len(cap_pool) else 150
    # Group spits by (component, date) and cap each group
    from collections import defaultdict
    day_groups = defaultdict(list)
    for i, e in enumerate(events):
        if e["reject_code"] != 0:
            day_key = e["file_dt"][:10]  # extract YYYY-MM-DD from ISO datetime
            day_groups[(e["component"], day_key)].append(i)
    for (comp, _day), indices in day_groups.items():
        unit_cost = float(COMPONENTS[comp][2])
        cap_limit = float(COMPONENT_SPIT_CAPS.get(comp, 150))
        max_allowed = int((cap_limit - 1e-9) / unit_cost) if unit_cost > 0 else 0
        if len(indices) > max_allowed:
            excess = len(indices) - max_allowed
            to_convert = random.sample(indices, excess)
            for idx in to_convert:
                events[idx]["reject_code"] = 0
    conn = sqlite3.connect(":memory:",check_same_thread=False)
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("CREATE TABLE logs(file_hash TEXT PRIMARY KEY,filename TEXT NOT NULL,file_dt TEXT,machine TEXT,board_name TEXT,mo TEXT,source_path TEXT,source_mtime REAL,source_size INTEGER,ingested_at TEXT NOT NULL)")
    conn.executemany("INSERT INTO logs VALUES(?,?,?,?,?,?,?,?,?,?)",[(l["file_hash"],l["filename"],l["file_dt"],l["machine"],l["board_name"],l["mo"],l["source_path"],l["source_mtime"],l["source_size"],l["ingested_at"]) for l in logs])
    conn.execute("CREATE TABLE events(id INTEGER PRIMARY KEY AUTOINCREMENT,file_hash TEXT NOT NULL,source_row_index INTEGER,component TEXT,description TEXT,location TEXT,board_name TEXT,mo TEXT,file_dt TEXT,machine TEXT,unit_cost REAL,cost REAL,reject_code INTEGER,feeder_no TEXT,slot_no TEXT)")
    conn.executemany("INSERT INTO events(file_hash,source_row_index,component,description,location,feeder_no,slot_no,board_name,mo,file_dt,machine,unit_cost,cost,reject_code) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)",[(e["file_hash"],e["source_row_index"],e["component"],e["description"],e["location"],e["feeder_no"],e["slot_no"],e["board_name"],e["mo"],e["file_dt"],e["machine"],e["unit_cost"],e["cost"],e["reject_code"]) for e in events])
    conn.commit()
    conn.execute("CREATE TABLE bom_items(bom_id INTEGER,component TEXT,unit_cost REAL)")
    for comp,(_,_,cost) in COMPONENTS.items(): conn.execute("INSERT INTO bom_items VALUES(?,?,?)",(1,comp.upper(),cost))
    conn.execute("CREATE TABLE bom_versions(bom_id INTEGER,bom_name TEXT,uploaded_at TEXT)")
    conn.execute("INSERT INTO bom_versions VALUES(1,'Demo Master BOM','2026-01-01 00:00:00')"); conn.commit()
    for c in ["file_dt","board_name","mo","machine","component","reject_code","feeder_no","slot_no"]:
        try: conn.execute(f"CREATE INDEX IF NOT EXISTS idx_events_{c} ON events({c})")
        except: pass
    for c in ["file_dt","board_name","mo","machine"]:
        try: conn.execute(f"CREATE INDEX IF NOT EXISTS idx_logs_{c} ON logs({c})")
        except: pass
    return conn,{"boards":sorted(set(l["board_name"] for l in logs)),"mos":sorted(set(l["mo"] for l in logs)),"machines":sorted(set(l["machine"] for l in logs)),"components":sorted(set(e["component"] for e in events))}

DEMO_CONN, DEMO_FILTERS = create_demo_database()
BOM_LOOKUP = {r[0].strip().upper():float(r[1]) for r in DEMO_CONN.execute("SELECT component,unit_cost FROM bom_items").fetchall()}
def get_bom_lookup(conn=None,selected_bom_ids=None): return BOM_LOOKUP
def list_boms(conn=None): return pd.read_sql_query("SELECT bom_id,bom_name,uploaded_at FROM bom_versions ORDER BY bom_id DESC",DEMO_CONN)

def _build_where(dt_start,dt_end,boards,mos,machines,components=None):
    w,p=[],[]
    if dt_start: w.append("(file_dt IS NOT NULL AND file_dt >= ?)"); p.append(dt_start.isoformat(sep=" "))
    if dt_end: w.append("(file_dt IS NOT NULL AND file_dt <= ?)"); p.append(dt_end.isoformat(sep=" "))
    if boards: w.append("board_name IN (%s)"%",".join(["?"]*len(boards))); p.extend(boards)
    if mos: w.append("mo IN (%s)"%",".join(["?"]*len(mos))); p.extend(mos)
    if machines: w.append("machine IN (%s)"%",".join(["?"]*len(machines))); p.extend(machines)
    if components: w.append("component IN (%s)"%",".join(["?"]*len(components))); p.extend(components)
    return w,p

def query_total_placement_cost(conn,dt_start,dt_end,boards,mos,machines,bl):
    w,p=_build_where(dt_start,dt_end,boards,mos,machines); w.append("reject_code=0")
    s="SELECT component,COUNT(*) AS n FROM events"
    if w: s+=" WHERE "+" AND ".join(w)
    s+=" GROUP BY component"
    df=pd.read_sql_query(s,conn,params=p)
    if df.empty: return 0.0
    df["n"]=pd.to_numeric(df["n"],errors="coerce").fillna(0.0)
    df["uc"]=df["component"].map(lambda c:float(bl.get(c.strip().upper(),0.0)))
    return float((df["n"]*df["uc"]).sum())

def query_successful_placements(conn,dt_start,dt_end,boards,mos,machines,components,bl):
    w,p=_build_where(dt_start,dt_end,boards,mos,machines,components); w.append("reject_code=0")
    s="SELECT component AS Component,COUNT(*) AS SuccessfulCount FROM events"
    if w: s+=" WHERE "+" AND ".join(w)
    s+=" GROUP BY component"
    return pd.read_sql_query(s,conn,params=p)

def query_summary(conn,dt_start,dt_end,boards,mos,machines,components,bl):
    w,p=_build_where(dt_start,dt_end,boards,mos,machines,components)
    rcs=sorted(REJECT_CODES); w.append("reject_code IN (%s)"%",".join(["?"]*len(rcs))); p.extend(rcs)
    g=pd.read_sql_query("SELECT component AS Component,description AS Description,machine AS Machine,reject_code AS RejectCode,COUNT(*) AS Spits FROM events" + (" WHERE "+" AND ".join(w) if w else "") + " GROUP BY component,description,machine,reject_code",conn,params=p)
    ec=["Component","Description","Machine","Spits","Rejection Rate %","Reject Codes","UnitCost","TotalCost","TotalPlacementCost","Loss % of Placement Value","cumulative percentage"]
    if g.empty: return pd.DataFrame(columns=ec)
    g["Spits"]=pd.to_numeric(g["Spits"],errors="coerce").fillna(0.0)
    def _fmt(ser): ss=ser.dropna(); vc=ss.value_counts() if not ss.empty else []; return ", ".join(f"{int(cnt)}x C{int(cd)}" for cd,cnt in vc.sort_values(ascending=False).items()) if not ss.empty else ""
    dd=(g.groupby(["Component","Description"],dropna=False)["Spits"].sum().reset_index().sort_values(["Component","Spits","Description"],ascending=[True,False,True]).drop_duplicates(subset=["Component"])[["Component","Description"]])
    md=(g.groupby(["Component","Machine"],dropna=False)["Spits"].sum().reset_index().sort_values(["Component","Spits","Machine"],ascending=[True,False,True]).drop_duplicates(subset=["Component"])[["Component","Machine"]])
    rdf=g.groupby(["Component","RejectCode"],dropna=False)["Spits"].sum().reset_index().sort_values(["Component","Spits","RejectCode"],ascending=[True,False,True])
    rt=rdf.groupby("Component",sort=False).apply(lambda pt: ", ".join(f"{int(cnt)}x C{int(cd)}" for cd,cnt in zip(pt["RejectCode"],pt["Spits"]))).reset_index(name="Reject Codes")
    sm=(g.groupby("Component",sort=False)["Spits"].sum().reset_index().merge(dd,on="Component",how="left").merge(md,on="Component",how="left").merge(rt,on="Component",how="left"))
    sd=query_successful_placements(conn,dt_start,dt_end,boards,mos,machines,components,bl)
    if not sd.empty:
        sm=sm.merge(sd.groupby("Component")["SuccessfulCount"].sum().reset_index(),on="Component",how="left")
        sm["SuccessfulCount"]=pd.to_numeric(sm.get("SuccessfulCount",0),errors="coerce").fillna(0).astype(int)
    else: sm["SuccessfulCount"]=0
    sm["Rejection Rate %"]=sm.apply(lambda r:(float(r["Spits"])/(float(r["Spits"])+float(r.get("SuccessfulCount",0)))*100.0) if (float(r["Spits"])+float(r.get("SuccessfulCount",0)))>0 else 0.0,axis=1)
    sm["UnitCost"]=sm["Component"].map(lambda c:float(bl.get(c.strip().upper(),0.0)))
    sm["TotalCost"]=sm["Spits"]*sm["UnitCost"]
    tpc=query_total_placement_cost(conn,dt_start,dt_end,boards,mos,machines,bl)
    sm["TotalPlacementCost"]=tpc
    sm["Loss % of Placement Value"]=(sm["TotalCost"]/sm["TotalPlacementCost"])*100
    sm=sm.sort_values("Spits",ascending=False,na_position="last")
    sm["cumulative percentage"]=sm["Loss % of Placement Value"].fillna(0.0).cumsum()
    return sm[ec]

def estimate_total_boards(conn,dt_start,dt_end,boards,mos,machines):
    w,p=_build_where(dt_start,dt_end,boards,mos,machines)
    df=pd.read_sql_query("SELECT COUNT(*) AS n FROM logs" + (" WHERE "+" AND ".join(w) if w else ""),conn,params=p)
    return float(df["n"].sum()) if not df.empty else 0.0

def machine_log_breakdown(conn,dt_start,dt_end,boards,mos,machines):
    w,p=_build_where(dt_start,dt_end,boards,mos,machines)
    return pd.read_sql_query("SELECT machine AS Machine,COUNT(*) AS BoardsRun FROM logs" + (" WHERE "+" AND ".join(w) if w else "") + " GROUP BY machine ORDER BY BoardsRun DESC",conn,params=p)

def query_events(conn,dt_start,dt_end,boards,mos,machines,components,bl):
    w,p=_build_where(dt_start,dt_end,boards,mos,machines,components)
    s="SELECT component AS Component,description AS Description,location AS Location,feeder_no AS Feeder,slot_no AS Slot,board_name AS Board,mo AS MO,file_dt AS FileDateTime,machine AS Machine,reject_code AS RejectCode FROM events"
    if w: s+=" WHERE "+" AND ".join(w)
    s+=" ORDER BY file_dt DESC"
    df=pd.read_sql_query(s,conn,params=p)
    if df.empty: df["UnitCost"],df["Cost"]=[],[]
    else: df["UnitCost"]=df["Component"].map(lambda c:float(bl.get(c.strip().upper(),0.0))); df["Cost"]=df["UnitCost"]
    return df

def query_repeated_locations(conn,dt_start,dt_end,boards,mos,machines,components,bl):
    w,p=_build_where(dt_start,dt_end,boards,mos,machines,components)
    rc=sorted(REJECT_CODES); w.append("reject_code IN (%s)"%",".join(["?"]*len(rc))); p.extend(rc)
    s="SELECT component AS Component,location AS Location,board_name AS Board,machine AS Machine,COUNT(*) AS Spits FROM events"
    if w: s+=" WHERE "+" AND ".join(w)
    s+=" GROUP BY component,location,board_name,machine HAVING COUNT(*)>1 ORDER BY COUNT(*) DESC"
    df=pd.read_sql_query(s,conn,params=p)
    if df.empty: return pd.DataFrame(columns=["Component","Location","Board","Machine","Spits","TotalCost","BoardsSeen","MultiBoard"])
    df["UnitCost"]=df["Component"].map(lambda c:float(bl.get(c.strip().upper(),0.0)))
    df["TotalCost"]=pd.to_numeric(df["Spits"],errors="coerce").fillna(0.0)*df["UnitCost"]
    df["BoardsSeen"],df["MultiBoard"]=1,"No"
    return df

def query_missing_costs(conn,dt_start,dt_end,boards,mos,machines,components,bl):
    w,p=_build_where(dt_start,dt_end,boards,mos,machines,components)
    rc=sorted(REJECT_CODES); w.append("reject_code IN (%s)"%",".join(["?"]*len(rc))); p.extend(rc)
    s="SELECT component AS Component,COUNT(*) AS [Spits (cost=0)] FROM events"
    if w: s+=" WHERE "+" AND ".join(w)
    s+=" GROUP BY component"
    df=pd.read_sql_query(s,conn,params=p)
    if df.empty: return pd.DataFrame(columns=["Component","Spits (cost=0)"])
    return df[df["Component"].map(lambda c:float(bl.get(c.strip().upper(),0.0)))==0.0].sort_values("Spits (cost=0)",ascending=False)

def query_feeder_slot_heatmap(conn,dt_start,dt_end,boards,mos,machines,components,bl,reject_codes=None):
    w,p=_build_where(dt_start,dt_end,boards,mos,machines,components)
    ac=sorted(REJECT_CODES if reject_codes is None else [int(c) for c in reject_codes])
    w.append("reject_code IN (%s)"%",".join(["?"]*len(ac))); p.extend(ac)
    s="SELECT machine AS Machine,feeder_no AS Feeder,slot_no AS Slot,component AS Component,description AS Description,Cost AS Cost FROM events"
    if w: s+=" WHERE "+" AND ".join(w)
    df=pd.read_sql_query(s,conn,params=p)
    if df.empty: return pd.DataFrame(columns=["Machine","Feeder","Slot","Spits","TotalCost","TopComponent","TopDescription"])
    df=df.dropna(subset=["Feeder","Slot"]); df=df[(df["Feeder"]!="")&(df["Slot"]!="")]
    grp=df.groupby(["Machine","Feeder","Slot"],dropna=False).agg(Spits=("Component","count"),TotalCost=("Cost","sum")).reset_index()
    tp=(df.groupby(["Machine","Feeder","Slot","Component","Description"],dropna=False).size().reset_index(name="Count").sort_values(["Machine","Feeder","Slot","Count"],ascending=[True,True,True,False]).drop_duplicates(subset=["Machine","Feeder","Slot"])[["Machine","Feeder","Slot","Component","Description"]].rename(columns={"Component":"TopComponent","Description":"TopDescription"}))
    return grp.merge(tp,on=["Machine","Feeder","Slot"],how="left")

def _label_sort_key(value):
    t="" if value is None else str(value).strip()
    if t=="": return (1,"")
    try: return (0,float(t))
    except: m=re.match(r"^\s*(\d+)",t); return (0,float(m.group(1)),t) if m else (1,t.lower())

def query_component_trend(conn,dt_start,dt_end,boards,mos,machines,components,component):
    w,p=_build_where(dt_start,dt_end,boards,mos,machines,components)
    rc=sorted(REJECT_CODES); w.append("reject_code IN (%s)"%",".join(["?"]*len(rc))); p.extend(rc)
    w.append("component = ?"); p.append(component)
    delta=dt_end-dt_start if dt_start and dt_end else timedelta(days=7)
    if delta<=timedelta(hours=6): be="strftime('%Y-%m-%d %H:',file_dt)||printf('%02d:00',(CAST(strftime('%M',file_dt) AS INTEGER)/30)*30)"
    elif delta<=timedelta(days=1): be="substr(file_dt,1,13)||':00:00'"
    elif delta<=timedelta(days=5): be="strftime('%Y-%m-%d ',file_dt)||printf('%02d:00:00',(CAST(strftime('%H',file_dt) AS INTEGER)/6)*6)"
    elif delta<=timedelta(days=45): be="substr(file_dt,1,10)||' 00:00:00'"
    else: be="date(file_dt,'-6 days','weekday 1')||' 00:00:00'"
    return pd.read_sql_query(f"SELECT {be} AS Bucket,COUNT(*) AS Spits FROM events" + (" WHERE "+" AND ".join(w) if w else "") + " GROUP BY Bucket ORDER BY Bucket",conn,params=p)

def query_success_ratio(conn,dt_start,dt_end,boards,mos,machines,components,bl):
    sm=query_summary(conn,dt_start,dt_end,boards,mos,machines,components,bl)
    sd=query_successful_placements(conn,dt_start,dt_end,boards,mos,machines,components,bl)
    ac=set()
    if not sm.empty: ac.update(sm["Component"].tolist())
    if not sd.empty: ac.update(sd["Component"].tolist())
    ratio=sm[["Component","Spits"]].copy() if not sm.empty else pd.DataFrame(columns=["Component","Spits"])
    ratio=ratio.merge(sd.groupby("Component",as_index=False)["SuccessfulCount"].sum(),on="Component",how="outer")
    if "SuccessfulCount" not in ratio.columns: ratio["SuccessfulCount"]=0
    if "Spits" not in ratio.columns: ratio["Spits"]=0
    ratio["SuccessfulCount"]=pd.to_numeric(ratio.get("SuccessfulCount",0),errors="coerce").fillna(0).astype(int)
    ratio["Spits"]=pd.to_numeric(ratio.get("Spits",0),errors="coerce").fillna(0).astype(int)
    ratio["SpitToSuccessPct"]=ratio.apply(lambda r:(float(r["Spits"])/float(r["SuccessfulCount"]))*100.0 if float(r["SuccessfulCount"])>0 else 0.0,axis=1)
    if ac:
        w,p=_build_where(dt_start,dt_end,boards,mos,machines,list(ac))
        dd=pd.read_sql_query("SELECT component AS Component,description AS Description FROM events" + (" WHERE "+" AND ".join(w) if w else "") + " GROUP BY component",conn,params=p)
        dd=dd[dd["Description"].notna()&(dd["Description"]!="")].drop_duplicates(subset=["Component"],keep="first")
        ratio=ratio.merge(dd,on="Component",how="left")
    if "Description" not in ratio.columns: ratio["Description"]=""
    ratio["Description"]=ratio["Description"].fillna("")
    return ratio.sort_values(["SpitToSuccessPct","Spits"],ascending=[False,False])

def query_successful_trend(conn,dt_start,dt_end,boards,mos,machines,components,comp):
    w,p=_build_where(dt_start,dt_end,boards,mos,machines,components)
    w.append("reject_code=0"); w.append("component = ?"); p.append(comp)
    delta=dt_end-dt_start if dt_start and dt_end else timedelta(days=7)
    if delta<=timedelta(days=1): be="substr(file_dt,1,13)||':00:00'"
    elif delta<=timedelta(days=5): be="strftime('%Y-%m-%d ',file_dt)||printf('%02d:00:00',(CAST(strftime('%H',file_dt) AS INTEGER)/6)*6)"
    else: be="substr(file_dt,1,10)||' 00:00:00'"
    return pd.read_sql_query(f"SELECT {be} AS Bucket,COUNT(*) AS SuccessCount FROM events" + (" WHERE "+" AND ".join(w) if w else "") + " GROUP BY Bucket ORDER BY Bucket",conn,params=p)

def query_spit_timeline(conn,dt_start,dt_end,boards,mos,machines,components):
    w,p=_build_where(dt_start,dt_end,boards,mos,machines,components)
    rc=sorted(REJECT_CODES); w.append("reject_code IN (%s)"%",".join(["?"]*len(rc))); p.extend(rc)
    delta=dt_end-dt_start if dt_start and dt_end else timedelta(days=7)
    if delta<=timedelta(days=1): be="substr(file_dt,1,13)||':00:00'"
    elif delta<=timedelta(days=5): be="strftime('%Y-%m-%d ',file_dt)||printf('%02d:00:00',(CAST(strftime('%H',file_dt) AS INTEGER)/6)*6)"
    else: be="substr(file_dt,1,10)||' 00:00:00'"
    return pd.read_sql_query(f"SELECT {be} AS Bucket,COUNT(*) AS Spits FROM events" + (" WHERE "+" AND ".join(w) if w else "") + " GROUP BY Bucket ORDER BY Bucket",conn,params=p)

def _wrap_datetime(dt):
    """Map any datetime into the 30-day demo data window (2026-05-11 to 2026-06-10).
    Recycles the same 30 days of synthetic data for any selected date."""
    anchor = date(2026, 5, 11)
    offset = (dt.date() - anchor).days % 30
    return datetime.combine(anchor + timedelta(days=offset), dt.time())

def _shift_query_context_to_previous_period(qctx):
    dt_start=datetime.fromisoformat(qctx["dt_start"]) if qctx.get("dt_start") else None
    dt_end=datetime.fromisoformat(qctx["dt_end"]) if qctx.get("dt_end") else None
    if dt_start is None or dt_end is None or dt_end<=dt_start: return None
    # Wrap dates into demo window so previous period aligns with available data
    dt_start = _wrap_datetime(dt_start)
    dt_end = _wrap_datetime(dt_end)
    window=dt_end-dt_start; prev_end=dt_start; prev_start=dt_start-window
    return {"dt_start":prev_start.isoformat(sep=" "),"dt_end":prev_end.isoformat(sep=" "),
            "boards":list(qctx.get("boards",[])),"mos":list(qctx.get("mos",[])),
            "machines":list(qctx.get("machines",[])),"components":list(qctx.get("components",[]))}

# PLOT HELPERS
def make_component_trend_figure(trend_df,component,bucket="day"):
    if trend_df.empty: return None
    fig=go.Figure(data=[go.Scatter(x=trend_df["Bucket"],y=trend_df["Spits"],mode="lines+markers",line=dict(color="#2196F3",width=2),marker=dict(size=6))])
    fig.update_layout(template="plotly_white",title=f"Spit Trend — {component}",height=360,margin=dict(l=20,r=20,t=50,b=20))
    return fig

def make_bucketed_count_figure(trend_df,title,y_label,color="#4e79a7"):
    if trend_df.empty: return None
    col=trend_df.columns[1] if len(trend_df.columns)>1 else trend_df.columns[0]
    fig=go.Figure(data=[go.Scatter(x=trend_df["Bucket"],y=trend_df[col],mode="lines+markers",line=dict(color=color,width=2),marker=dict(size=6))])
    fig.update_layout(template="plotly_white",title=title,height=380,margin=dict(l=20,r=20,t=50,b=20))
    return fig

def build_pareto_data(summary_df,metric):
    if summary_df.empty: return pd.DataFrame()
    mc="TotalCost" if metric=="Cost" else "Spits"
    df=summary_df[["Component",mc]].copy().rename(columns={mc:"Value"}).sort_values("Value",ascending=False).reset_index(drop=True)
    total=float(df["Value"].sum())
    df["CumulativePercent"]=(df["Value"].cumsum()/total*100.0) if total>0 else 0.0
    return df

def make_pareto_figure(pareto_df,metric,top_n=20):
    if pareto_df.empty: return None
    d=pareto_df.head(top_n); vl="Total Cost" if metric=="Cost" else "Spits"
    fig=go.Figure()
    fig.add_bar(x=d["Component"],y=d["Value"],name=vl,marker_color="#2196F3")
    fig.add_scatter(x=d["Component"],y=d["CumulativePercent"],name="Cumulative %",mode="lines+markers",yaxis="y2",line=dict(color="#E65100",width=2))
    fig.update_layout(template="plotly_white",title=f"Pareto by {vl}",yaxis2=dict(title="Cumulative %",overlaying="y",side="right",range=[0,100]),height=380,margin=dict(l=20,r=20,t=50,b=20))
    fig.update_xaxes(tickangle=-35)
    return fig

def make_feeder_slot_heatmap_figure_with_options(heatmap_df,cost_based,color_scale,clip_percentile,title,height_override=None):
    if heatmap_df.empty: return None
    feeders=sorted(heatmap_df["Feeder"].dropna().unique().tolist(),key=_label_sort_key)
    slots=sorted(heatmap_df["Slot"].dropna().unique().tolist(),key=_label_sort_key)
    if not feeders or not slots: return None
    cm=pd.DataFrame(0,index=feeders,columns=slots)
    cst=pd.DataFrame(0.0,index=feeders,columns=slots)
    comp_mat=pd.DataFrame("",index=feeders,columns=slots)
    desc_mat=pd.DataFrame("",index=feeders,columns=slots)
    for _,row in heatmap_df.iterrows():
        if row["Feeder"] in cm.index and row["Slot"] in cm.columns:
            cm.at[row["Feeder"],row["Slot"]]=int(row["Spits"])
            cst.at[row["Feeder"],row["Slot"]]=float(row["TotalCost"])
            comp_mat.at[row["Feeder"],row["Slot"]]=str(row.get("TopComponent","") or "")
            desc_mat.at[row["Feeder"],row["Slot"]]=str(row.get("TopDescription","") or "")
    z=(cst if cost_based else cm).values.astype(float)
    ct="Total Cost" if cost_based else "Spit Count"
    zmax=None
    if clip_percentile<100:
        try: zv=z.flatten(); zv=zv[zv>0]; zmax=float(np.percentile(zv,clip_percentile)) if len(zv)>0 and clip_percentile<100 else None
        except: pass
    cd=[[[int(cm.iat[i,j]),float(cst.iat[i,j]),str(comp_mat.iat[i,j]),str(desc_mat.iat[i,j])] for j in range(len(slots))] for i in range(len(feeders))]
    fig=go.Figure(data=go.Heatmap(z=z,x=slots,y=feeders,colorscale=color_scale,zmin=0,zmax=zmax,customdata=cd,hovertemplate="Feeder %{y}<br>Slot %{x}<br>Spit count: %{customdata[0]}<br>Total cost: £%{customdata[1]:,.2f}<br>Top component: %{customdata[2]}<br>Description: %{customdata[3]}<extra></extra>",colorbar=dict(title=ct)))
    cell_h=max(42,min(60,800//max(1,len(feeders))))
    fig.update_layout(margin=dict(l=20,r=20,t=50,b=20),height=height_override or max(480,cell_h*len(feeders)),template="plotly_white",title=title,clickmode="event+select")
    fig.update_xaxes(title_text="Slot No.",side="top",tickmode="array",tickvals=slots,ticktext=[str(s) for s in slots]); fig.update_yaxes(title_text="Feeder No.",autorange="reversed")
    return fig

def extract_selected_heatmap_cell(plot_state):
    if plot_state is None: return None,None
    if isinstance(plot_state,dict): pts=plot_state.get("selection",{}).get("points",[])
    else: pts=getattr(getattr(plot_state,"selection",None),"points",[])
    if not pts: return None,None
    p=pts[0]
    return (str(p.get("y","")),str(p.get("x",""))) if isinstance(p,dict) else (str(getattr(p,"y","")),str(getattr(p,"x","")))

def _aggregate_heatmap_scope(heatmap_df,scope,machine=None):
    if heatmap_df.empty: return heatmap_df
    if scope=="Single machine" and machine: return heatmap_df[heatmap_df["Machine"]==machine].drop(columns=["Machine"],errors="ignore")
    combined = heatmap_df.groupby(["Feeder","Slot"],dropna=False).agg(Spits=("Spits","sum"),TotalCost=("TotalCost","sum")).reset_index()
    top = (heatmap_df.groupby(["Feeder","Slot","TopComponent","TopDescription"],dropna=False)["Spits"].sum().reset_index()
           .sort_values(["Feeder","Slot","Spits",],ascending=[True,True,False])
           .drop_duplicates(subset=["Feeder","Slot"])
           [["Feeder","Slot","TopComponent","TopDescription"]])
    return combined.merge(top,on=["Feeder","Slot"],how="left")

def compute_selection_overview(conn,dt_start,dt_end,boards,mos,machines,components):
    bl=get_bom_lookup()
    sm=query_summary(conn,dt_start,dt_end,boards,mos,machines,components,bl)
    hm=query_feeder_slot_heatmap(conn,dt_start,dt_end,boards,mos,machines,components,bl)
    rp=query_repeated_locations(conn,dt_start,dt_end,boards,mos,machines,components,bl)
    ms=query_missing_costs(conn,dt_start,dt_end,boards,mos,machines,components,bl)
    ts=int(sm["Spits"].sum()) if not sm.empty else 0
    tl=float(sm["TotalCost"].sum()) if not sm.empty else 0.0
    br=float(estimate_total_boards(conn,dt_start,dt_end,boards,mos,machines))
    bmd=machine_log_breakdown(conn,dt_start,dt_end,boards,mos,machines)
    wc,wcc,wcs="-",0.0,0; hsc,hscs="-",0
    if not sm.empty:
        top=sm.sort_values(["TotalCost","Spits"],ascending=[False,False]).iloc[0]
        wc=str(top.get("Component","-")); wcc=float(top.get("TotalCost",0.0)); wcs=int(top.get("Spits",0))
        t2=sm.sort_values(["Spits","TotalCost"],ascending=[False,False]).iloc[0]
        hsc=str(t2.get("Component","-")); hscs=int(t2.get("Spits",0))
    wh,whc,whs="-",0.0,0
    if not hm.empty:
        top=hm.sort_values(["Spits","TotalCost"],ascending=[False,False]).iloc[0]
        wh=f"{top.get('Machine','-')}: F{top.get('Feeder','-')} / S{top.get('Slot','-')}"
        whc=float(top.get("TotalCost",0.0)); whs=int(top.get("Spits",0))
    wm,wms="-",0
    if not sm.empty and "Machine" in sm.columns:
        mr=sm.groupby("Machine")["Spits"].sum().reset_index().sort_values("Spits",ascending=False)
        if not mr.empty: wm=str(mr.iloc[0]["Machine"]); wms=int(mr.iloc[0]["Spits"])
    ar=[]
    if wc!="-": ar.append({"Priority":"High","Action":f"Review component {wc}","Reason":f"Cost driver: £{wcc:,.2f}"})
    if hsc!="-": ar.append({"Priority":"High","Action":f"Check component {hsc}","Reason":f"Spit-count driver: {hscs} spits"})
    if wm!="-": ar.append({"Priority":"Medium","Action":f"Check machine {wm}","Reason":f"Spit driver: {wms} spits"})
    if not ms.empty: ar.append({"Priority":"Low","Action":"Fill missing BOM costs","Reason":f"{len(ms)} components have zero cost"})
    ar.append({"Priority":"Low","Action":"Review BOM cost accuracy","Reason":"Ensure all component costs are up-to-date in the master BOM"})
    return {"total_spits":ts,"total_loss":tl,"boards_run":br,"boards_by_machine_df":bmd,
            "worst_component":wc,"worst_component_cost":wcc,"worst_component_spits":wcs,
            "highest_spit_component":hsc,"highest_spit_component_spits":hscs,
            "worst_hotspot":wh,"worst_hotspot_cost":whc,"worst_machine":wm,"worst_machine_spits":wms,
            "action_df":pd.DataFrame(ar,columns=["Priority","Action","Reason"])}

# APP UI

logo_path = os.path.join(os.path.dirname(__file__), "pyline_logo.png") if "__file__" in dir() else "pyline_logo.png"
logo_html = ''
if os.path.exists(logo_path):
    import base64 as _b64
    with open(logo_path, "rb") as _lf:
        _logo_b64 = _b64.b64encode(_lf.read()).decode()
    logo_html = f'<img src="data:image/png;base64,{_logo_b64}" style="height:38px;">'

st.markdown(f"""<div class="proline-header">
    {logo_html}
    <div class="brand-text"><h1>PROLINE ANALYTICS</h1><div class="subtitle">PCBA Spit Analytics — Real-Time Manufacturing Intelligence</div></div>
</div>""",unsafe_allow_html=True)

default_start=datetime.combine(date(2026,6,9),time(0,0,0))
if "end_date_default_set" not in st.session_state:
    st.session_state["end_date"]=date(2026,6,9); st.session_state["end_time"]=time(22,30); st.session_state["end_date_default_set"]=True

compare_mode=st.checkbox("Compare Periods",key="compare_mode",help="Select two different date ranges to compare")

if st.session_state.get("compare_mode",False):
    st.caption("Period A")
    fc1a,fc2a,fc3a,fc4a=st.columns([0.9,0.7,0.9,0.7])
    with fc1a: start_date=st.date_input("P1 Start",value=date(2026,6,8),key="start_date_a")
    with fc2a: start_time=st.time_input("P1 Start time",value=time(0,0),key="start_time_a")
    with fc3a: end_date=st.date_input("P1 End",value=date(2026,6,8),key="end_date_a")
    with fc4a: end_time=st.time_input("P1 End time",value=time(21,15),key="end_time_a")
    st.caption("Period B")
    fc1b,fc2b,fc3b,fc4b,fc5b=st.columns([0.9,0.7,0.9,0.7,1.4])
    with fc1b: start_date_b=st.date_input("P2 Start",value=date(2026,6,7),key="start_date_b")
    with fc2b: start_time_b=st.time_input("P2 Start time",value=time(0,0),key="start_time_b")
    with fc3b: end_date_b=st.date_input("P2 End",value=date(2026,6,7),key="end_date_b")
    with fc4b: end_time_b=st.time_input("P2 End time",value=time(21,15),key="end_time_b")
    with fc5b: run_query=st.button("🔍 Run Comparison",type="primary",use_container_width=True)
    dt_start=datetime.combine(start_date,start_time); dt_end=datetime.combine(end_date,end_time)
    dt_start_b=datetime.combine(start_date_b,start_time_b); dt_end_b=datetime.combine(end_date_b,end_time_b)
    compare_period_b={"dt_start":dt_start_b,"dt_end":dt_end_b}
else:
    fc1,fc2,fc3,fc4,fc5=st.columns([0.9,0.7,0.9,0.7,1.4])
    with fc1: start_date=st.date_input("Start date",value=default_start.date(),key="start_date")
    with fc2: start_time=st.time_input("Start time",value=default_start.time(),key="start_time")
    with fc3: end_date=st.date_input("End date",key="end_date")
    with fc4: end_time=st.time_input("End time",key="end_time")
    with fc5: run_query=st.button("🔍 Run Query",type="primary",use_container_width=True)
    dt_start=datetime.combine(start_date,start_time); dt_end=datetime.combine(end_date,end_time)
    compare_period_b=None

with st.expander("⚙️ Advanced Filters",expanded=False):
    af0,af1,af2=st.columns([1,1,1])
    with af0: boards_sel=st.multiselect("Board Name",DEMO_FILTERS["boards"],default=[],key="boards_sel")
    with af1: mos_sel=st.multiselect("MO",DEMO_FILTERS["mos"],default=[],key="mos_sel")
    with af2: machines_sel=st.multiselect("Machine",DEMO_FILTERS["machines"],default=[],key="machines_sel")
    components_sel=st.multiselect("Component",DEMO_FILTERS["components"],default=[],key="components_sel")

if "has_results" not in st.session_state: st.session_state.has_results=False
if "payload" not in st.session_state: st.session_state.payload=None
if run_query:
    st.session_state.payload={"dt_start":dt_start.isoformat(),"dt_end":dt_end.isoformat(),"boards":boards_sel,"mos":mos_sel,"machines":machines_sel,"components":components_sel}
    st.session_state.has_results=True
elif not st.session_state.has_results:
    st.session_state.payload={"dt_start":dt_start.isoformat(),"dt_end":dt_end.isoformat(),"boards":[],"mos":[],"machines":[],"components":[]}

payload=st.session_state.payload
dt_start=datetime.fromisoformat(payload["dt_start"]); dt_end=datetime.fromisoformat(payload["dt_end"])
boards_sel=payload.get("boards",[]); mos_sel=payload.get("mos",[])
machines_sel=payload.get("machines",[]); components_sel=payload.get("components",[])
# Wrap dates into the 30-day demo data window (2026-05-11 to 2026-06-10)
dt_start = _wrap_datetime(dt_start)
dt_end = _wrap_datetime(dt_end)

st.caption("C2=Failed vision before electrical | C3=Failed vision after electrical | C4=Failed electrical test | C5=Component lost | C6=Not picked up by machine | C7=Failed vision before pickup")
vo=["Summary","Successful Placements","Spit Events","Heatmap of Trolleys and Slots","Repeated Locations","Missing BOM Costs"]
if st.session_state.get("compare_mode",False) and compare_period_b is not None: vo=["📊 Period Comparison"]+vo
st.markdown("""<style>
    #av-wrapper div[data-testid="stSelectbox"] label p { font-size:16px !important;font-weight:700 !important;color:#1B3A5C !important; }
</style><div id="av-wrapper" style="margin:0;padding:0;">""", unsafe_allow_html=True)
view=st.selectbox("Analysis view",vo,index=0)
st.markdown("</div>", unsafe_allow_html=True)
bl=get_bom_lookup()

prev_qctx=_shift_query_context_to_previous_period(st.session_state.payload)

# PERIOD COMPARISON VIEW
if view=="📊 Period Comparison" and compare_period_b is not None:
    oa=compute_selection_overview(DEMO_CONN,dt_start,dt_end,boards_sel,mos_sel,machines_sel,components_sel)
    ob=compute_selection_overview(DEMO_CONN,dt_start_b,dt_end_b,boards_sel,mos_sel,machines_sel,components_sel)
    la=float(oa.get('total_loss',0)); lb=float(ob.get('total_loss',0))
    sa=int(oa.get('total_spits',0)); sb=int(ob.get('total_spits',0))
    ba=int(oa.get('boards_run',0)); bb=int(ob.get('boards_run',0))
    ld=la-lb; lp=(ld/lb*100) if lb>0 else 0; iw=ld>0
    hc="worse" if iw else "better"; hp="higher" if iw else "lower"; he="🔴" if iw else "🟢"; hd="+" if iw else ""
    st.markdown(f"""<div class="comp-hero {hc}"><div class="hero-label">{he} COSTS IN PERIOD A ARE {hp.upper()}</div><div class="hero-value">{hd}£{abs(ld):,.0f}</div><div class="hero-detail">Period A Total Loss: <strong>£{la:,.2f}</strong> | Period B Total Loss: <strong>£{lb:,.2f}</strong> | <span style="color:{'#C62828' if iw else '#2E7D32'};font-weight:700;">{'▲' if iw else '▼'} {abs(lp):.1f}% {'increase' if iw else 'decrease'}</span></div></div>""",unsafe_allow_html=True)
    sd=sa-sb; bd=ba-bb
    st.markdown(f"""<div class="stat-row">
        <div class="stat-card"><div class="stat-label">Total Spits (A vs B)</div><div class="stat-num">{sa:,} / {sb:,}</div><div class="stat-sub">{'+' if sd>0 else ''}{sd:,} Δ</div></div>
        <div class="stat-card"><div class="stat-label">Boards Run (A vs B)</div><div class="stat-num">{ba:,} / {bb:,}</div><div class="stat-sub">{'+' if bd>0 else ''}{bd:,} Δ</div></div>
        <div class="stat-card"><div class="stat-label">Loss Per Board (A vs B)</div><div class="stat-num">£{(la/max(1,ba)):,.2f} / £{(lb/max(1,bb)):,.2f}</div><div class="stat-sub">avg per board</div></div>
        <div class="stat-card"><div class="stat-label">Worst Component (A)</div><div class="stat-num" style="font-size:18px;">{oa.get('worst_component','-')}</div><div class="stat-sub">£{float(oa.get('worst_component_cost',0)):,.2f} loss</div></div>""",unsafe_allow_html=True)
    sma=query_summary(DEMO_CONN,dt_start,dt_end,boards_sel,mos_sel,machines_sel,components_sel,bl)
    smb=query_summary(DEMO_CONN,dt_start_b,dt_end_b,boards_sel,mos_sel,machines_sel,components_sel,bl)
    cl,cr=st.columns([1,1])
    with cl:
        if not sma.empty and not smb.empty:
            aa=sma[["Component","Spits","TotalCost"]].copy(); aa.columns=["Component","Spits (A)","Cost (A)"]
            ab=smb[["Component","Spits","TotalCost"]].copy(); ab.columns=["Component","Spits (B)","Cost (B)"]
            cf=aa.merge(ab,on="Component",how="outer").fillna(0)
            cf["Δ Spits"]=cf["Spits (A)"].astype(int)-cf["Spits (B)"].astype(int); cf["Δ Cost"]=cf["Cost (A)"]-cf["Cost (B)"]
            cf=cf.sort_values("Δ Cost",ascending=False).head(12)
            def _cs(row): d=row.get("Δ Cost",0) if not isinstance(row,int) else 0; return(["background-color:#FFF5F5;color:#C62828;"]*len(row)) if d>0 else(["background-color:#F5FDF5;color:#2E7D32;"]*len(row)) if d<0 else [""]*len(row)
            st.subheader("Component Changes")
            st.dataframe(cf.style.apply(_cs,axis=1),use_container_width=True,hide_index=True,column_config={"Spits (A)":st.column_config.NumberColumn("Spit A",format="%d"),"Spits (B)":st.column_config.NumberColumn("Spit B",format="%d"),"Δ Spits":st.column_config.NumberColumn("Δ Spit",format="%+d"),"Cost (A)":st.column_config.NumberColumn("£ A",format="£%.2f"),"Cost (B)":st.column_config.NumberColumn("£ B",format="£%.2f"),"Δ Cost":st.column_config.NumberColumn("Δ Cost",format="£%+.2f")})
    with cr:
        ma=machine_log_breakdown(DEMO_CONN,dt_start,dt_end,boards_sel,mos_sel,machines_sel)
        mb=machine_log_breakdown(DEMO_CONN,dt_start_b,dt_end_b,boards_sel,mos_sel,machines_sel)
        if not ma.empty and not mb.empty:
            xa=ma.rename(columns={"BoardsRun":"Boards (A)"}); xb=mb.rename(columns={"BoardsRun":"Boards (B)"})
            mc=xa.merge(xb,on="Machine",how="outer").fillna(0); mc["Δ Boards"]=mc["Boards (A)"].astype(int)-mc["Boards (B)"].astype(int)
            mc=mc.sort_values("Δ Boards",ascending=False).head(12)
            st.subheader("Machine Changes")
            st.dataframe(mc,use_container_width=True,hide_index=True,column_config={"Boards (A)":st.column_config.NumberColumn("Boards A",format="%d"),"Boards (B)":st.column_config.NumberColumn("Boards B",format="%d"),"Δ Boards":st.column_config.NumberColumn("Δ",format="%+d")})
    tla=query_spit_timeline(DEMO_CONN,dt_start,dt_end,boards_sel,mos_sel,machines_sel,components_sel)
    tlb=query_spit_timeline(DEMO_CONN,dt_start_b,dt_end_b,boards_sel,mos_sel,machines_sel,components_sel)
    if not tla.empty and not tlb.empty:
        ta=tla.rename(columns={"Spits":"Period A"}); tb=tlb.rename(columns={"Spits":"Period B"})
        mg=ta.merge(tb,on="Bucket",how="outer").fillna(0).sort_values("Bucket"); mg["Period A"]=mg["Period A"].astype(int); mg["Period B"]=mg["Period B"].astype(int)
        tf=go.Figure()
        tf.add_scatter(x=mg["Bucket"],y=mg["Period A"],name="Period A",mode="lines+markers",line=dict(color="#2196F3",width=2.5))
        tf.add_scatter(x=mg["Bucket"],y=mg["Period B"],name="Period B",mode="lines+markers",line=dict(color="#9E9E9E",width=2.5,dash="dash"))
        tf.update_layout(template="plotly_white",title="Spit Timeline: Period A vs Period B",height=380,margin=dict(l=20,r=20,t=50,b=20)); st.plotly_chart(tf,use_container_width=True)
    if not sma.empty and not smb.empty and not cf.empty:
        pdata=cf.head(10); cfig=go.Figure()
        cfig.add_bar(x=pdata["Component"],y=pdata["Cost (A)"],name="Period A",marker_color="#2196F3")
        cfig.add_bar(x=pdata["Component"],y=pdata["Cost (B)"],name="Period B",marker_color="#BDBDBD")
        cfig.update_layout(template="plotly_white",title="Cost: Period A vs Period B",barmode="group",height=400,margin=dict(l=20,r=20,t=50,b=20)); cfig.update_xaxes(tickangle=-35); st.plotly_chart(cfig,use_container_width=True)
    st.stop()

# NORMAL VIEWS
if view=="Summary":
    overview=compute_selection_overview(DEMO_CONN,dt_start,dt_end,boards_sel,mos_sel,machines_sel,components_sel)
    prev_overview={}
    if prev_qctx:
        pds=datetime.fromisoformat(prev_qctx["dt_start"]) if prev_qctx.get("dt_start") else None
        pde=datetime.fromisoformat(prev_qctx["dt_end"]) if prev_qctx.get("dt_end") else None
        prev_overview=compute_selection_overview(DEMO_CONN,pds,pde,boards_sel,mos_sel,machines_sel,components_sel) if pds and pde else {}
    if overview:
        st.subheader("Selection Snapshot")
        def _trend_icon(dv):
            if dv is None: return ""
            if dv>0: return "▲"
            if dv<0: return "▼"
            return "—"
        def _metric_delta(cv,pv,fmt):
            if pv is None: return None
            try: dv=cv-pv; return f"{_trend_icon(dv)} {fmt(dv)}"
            except: return None
        ts=int(overview.get("total_spits",0)); tl=float(overview.get("total_loss",0))
        wc=str(overview.get("worst_component","-")); br=int(overview.get("boards_run",0))
        pts=prev_overview.get("total_spits") if prev_overview else None
        ptl=prev_overview.get("total_loss") if prev_overview else None
        sd_text=_metric_delta(ts,pts,lambda v:f"{int(v):+d} vs prev") if pts is not None else None
        ld_text=_metric_delta(tl,ptl,lambda v:f"{v:+,.2f} vs prev") if ptl is not None else None
        ts_sev="bad" if pts is not None and ts>pts else ("good" if pts is not None and ts<pts else "")
        tl_sev="bad" if ptl is not None and tl>ptl else ("good" if ptl is not None and tl<ptl else "")
        def _kpi(label,value,unit,severity="",delta_text=""):
            sc=severity if severity in("bad","good","warn") else "neutral"
            dot='<span class="status-dot red"></span>' if severity=="bad" else '<span class="status-dot green"></span>' if severity=="good" else '<span class="status-dot orange"></span>' if severity=="warn" else '<span class="status-dot blue"></span>'
            dh=f'<div class="kpi-delta">{delta_text}</div>' if delta_text else ""
            st.html(f"""<div class="kpi-card {sc}"><div class="kpi-label">{dot}{label}</div><div class="kpi-value">{value}<span class="kpi-unit"> {unit}</span></div>{dh}</div>""")
        k1,k2,k3,k4=st.columns(4)
        with k1: _kpi("Total Spits",f"{ts:,}","count",ts_sev,sd_text or "")
        with k2: _kpi("Total Loss",f"£{tl:,.2f}","currency",tl_sev,ld_text or "")
        with k3: _kpi("Boards Run",f"{br:,}","boards")
        with k4: _kpi("Worst Component",wc,f"£{overview.get('worst_component_cost',0):,.2f} loss","warn")
        mc1,mc2=st.columns([1.1,1.4])
        with mc1:
            st.caption(f"Worst hotspot (cost): {overview.get('worst_hotspot','-')} | £{overview.get('worst_hotspot_cost',0):,.2f} loss")
            bbm=overview.get("boards_by_machine_df",pd.DataFrame())
            if not bbm.empty: st.caption("Boards by machine"); st.dataframe(bbm,use_container_width=True,hide_index=True,column_config={"BoardsRun":st.column_config.NumberColumn("Boards",format="%d")})
        with mc2:
            adf=overview.get("action_df",pd.DataFrame())
            if not adf.empty:
                def _pc(v):
                    if v=="High": return "background-color:#fef0f0;color:#c62828;font-weight:600;border-left:4px solid #e15759;padding-left:8px;"
                    if v=="Medium": return "background-color:#fef8f0;color:#e65100;font-weight:600;border-left:4px solid #f28e2b;padding-left:8px;"
                    if v=="Low": return "background-color:#f0f7f0;color:#2e7d32;font-weight:600;border-left:4px solid #59a14f;padding-left:8px;"
                    return ""
                st.dataframe(adf.style.map(_pc,subset=["Priority"]),use_container_width=True,hide_index=True)
            else: st.info("No action items.")
    summary_df=query_summary(DEMO_CONN,dt_start,dt_end,boards_sel,mos_sel,machines_sel,components_sel,bl)
    sdd=summary_df.drop(columns=["TotalPlacementCost","SuccessfulCount"],errors="ignore").reset_index(drop=True)
    pcb=st.toggle("Cost-based Pareto",value=False,key="spc")
    pdf=build_pareto_data(summary_df,"Cost" if pcb else "Spit")
    pf=make_pareto_figure(pdf,"Cost" if pcb else "Spit")
    if pf: st.plotly_chart(pf,use_container_width=True)
    if not sdd.empty:
        st.dataframe(sdd,use_container_width=True,hide_index=True,key="sct",on_select="rerun",selection_mode="single-row",
            column_config={"Spits":st.column_config.NumberColumn("Spits",format="%d"),"Rejection Rate %":st.column_config.NumberColumn("Rejection Rate %",format="%.2f%%"),
                "UnitCost":st.column_config.NumberColumn("Unit Cost",format="%.2f"),"TotalCost":st.column_config.NumberColumn("Total Cost",format="%.2f"),
                "Loss % of Placement Value":st.column_config.NumberColumn("Loss % of Placement Value",format="%.2f"),"cumulative percentage":st.column_config.NumberColumn("Cumulative %",format="%.2f")})
        comps=sdd["Component"].astype(str).tolist()
        if comps:
            summ_state=st.session_state.get("sct")
            if summ_state and summ_state.get("selection") and summ_state["selection"].get("rows"):
                sel_idx=summ_state["selection"]["rows"][0]
                sel_comp=str(sdd.iloc[sel_idx]["Component"]) if sel_idx<len(sdd) else comps[0]
            else: sel_comp=comps[0]
            if sel_comp not in comps: sel_comp=comps[0]
            sc=st.selectbox("Selected component",options=comps,index=comps.index(sel_comp),key="scp")
            td=query_component_trend(DEMO_CONN,dt_start,dt_end,boards_sel,mos_sel,machines_sel,components_sel,sc)
            if not td.empty:
                tf=make_component_trend_figure(td,sc)
                if tf: st.plotly_chart(tf,use_container_width=True)

elif view=="Successful Placements":
    rdf=query_success_ratio(DEMO_CONN,dt_start,dt_end,boards_sel,mos_sel,machines_sel,components_sel,bl)
    if not rdf.empty:
        st.subheader("Spit to Success % by Component")
        rt=rdf[["Component","Description","SuccessfulCount","Spits","SpitToSuccessPct"]].sort_values(["SpitToSuccessPct","Spits"],ascending=[False,False]).reset_index(drop=True)
        st.dataframe(rt,use_container_width=True,hide_index=True,column_config={"SuccessfulCount":st.column_config.NumberColumn("Successful Placements",format="%d"),"Spits":st.column_config.NumberColumn("Spits",format="%d"),"SpitToSuccessPct":st.column_config.NumberColumn("Spit / Success %",format="%.2f%%")})
        comps2=rt["Component"].astype(str).tolist()
        if comps2:
            src=st.selectbox("Placement trend component",options=comps2,index=0,key="sprc")
            std=query_successful_trend(DEMO_CONN,dt_start,dt_end,boards_sel,mos_sel,machines_sel,components_sel,src)
            if not std.empty:
                sf=make_bucketed_count_figure(std,f"Successful Placements Over Time — {src}","Successful Placements","#2E7D32")
                if sf: st.plotly_chart(sf,use_container_width=True)
    st.subheader("Successful Placements Table")
    spd=query_successful_placements(DEMO_CONN,dt_start,dt_end,boards_sel,mos_sel,machines_sel,components_sel,bl)
    tpc_val_ui=query_total_placement_cost(DEMO_CONN,dt_start,dt_end,boards_sel,mos_sel,machines_sel,bl)
    spd["UnitCost"]=spd["Component"].map(lambda c:float(bl.get(c.strip().upper(),0.0))); spd["TotalCost"]=spd["SuccessfulCount"]*spd["UnitCost"]
    spd["TotalPlacementCost"]=tpc_val_ui
    cols_ordered = [c for c in spd.columns if c != "TotalPlacementCost"] + ["TotalPlacementCost"]
    st.dataframe(spd[cols_ordered],use_container_width=True)

elif view=="Spit Events":
    edf=query_events(DEMO_CONN,dt_start,dt_end,boards_sel,mos_sel,machines_sel,components_sel,bl)
    edf=edf[edf["RejectCode"].isin(REJECT_CODES)].reset_index(drop=True)
    tld=query_spit_timeline(DEMO_CONN,dt_start,dt_end,boards_sel,mos_sel,machines_sel,components_sel)
    if not tld.empty:
        tf=make_bucketed_count_figure(tld,"Spit Events Over Time","Spits","#C62828")
        if tf: st.plotly_chart(tf,use_container_width=True)
    st.dataframe(edf,use_container_width=True)

elif view=="Heatmap of Trolleys and Slots":
    hdf=query_feeder_slot_heatmap(DEMO_CONN,dt_start,dt_end,boards_sel,mos_sel,machines_sel,components_sel,bl)
    cb=st.toggle("Cost-based heatmap",value=False,key="hcb")
    if hdf.empty: st.info("No feeder/slot spit events found for the selected filters.")
    else:
        mo=sorted(hdf["Machine"].dropna().unique().tolist())
        scope=st.radio("Machine scope",["Combined selected machines","Single machine"],horizontal=True,key="hsc")
        cc1,cc2,cc3=st.columns([1,1,1])
        with cc1: cs=st.selectbox("Color palette",["YlOrRd","Viridis","Cividis","Blues","Reds"],index=0,key="hcs")
        with cc2: cp=st.slider("Color clipping percentile",90,100,100,1,key="hcp")
        with cc3: tn=st.number_input("Top hotspots",5,100,20,5,key="htn")
        sm=st.selectbox("Machine",options=mo,key="hsm") if scope=="Single machine" and len(mo)>1 else None
        if scope=="Single machine" and sm:
            scoped=hdf[hdf["Machine"]==sm].copy(); pdf=scoped.drop(columns=["Machine"],errors="ignore"); t=f"Heatmap — {sm}"
        else: scoped=hdf; pdf=_aggregate_heatmap_scope(hdf,"Combined"); t="Combined Heatmap (All Machines)"
        f=make_feeder_slot_heatmap_figure_with_options(pdf,cb,cs,cp,t)
        if f: st.plotly_chart(f,use_container_width=True,key=f"h_{scope}_{sm or 'all'}")
        else: st.info("No feeder/slot combinations available.")
        ws=scoped.sort_values(["Spits","TotalCost"],ascending=[False,False]).iloc[0]
        wc2=scoped.sort_values(["TotalCost","Spits"],ascending=[False,False]).iloc[0]
        uq=int(scoped[["Machine","Feeder","Slot"]].drop_duplicates().shape[0])
        k1,k2,k3=st.columns(3)
        with k1: st.metric("Worst by Spits",f"M{ws.get('Machine','')}: F{ws['Feeder']}/S{ws['Slot']}",delta=f"{int(ws['Spits'])} spits")
        with k2: st.metric("Worst by Cost",f"M{wc2.get('Machine','')}: F{wc2['Feeder']}/S{wc2['Slot']}",delta=f"£{float(wc2['TotalCost']):,.2f}")
        with k3: st.metric("Unique Hotspots",f"{uq}")
        ddf=scoped.copy(); ddf=ddf.sort_values("TotalCost" if cb else "Spits",ascending=False).head(tn)
        ddf=ddf.rename(columns={"TopComponent":"Component","TopDescription":"Description"})
        cols=[c for c in ["Machine","Feeder","Slot","Component","Description","Spits","TotalCost"] if c in ddf.columns]
        ddf=ddf[cols]
        st.subheader("Hotspot Table")
        st.dataframe(ddf,use_container_width=True,hide_index=True,column_config={"Spits":st.column_config.NumberColumn("Spits",format="%d"),"TotalCost":st.column_config.NumberColumn("Total Cost",format="£%.2f")})

elif view=="Repeated Locations":
    st.dataframe(query_repeated_locations(DEMO_CONN,dt_start,dt_end,boards_sel,mos_sel,machines_sel,components_sel,bl),use_container_width=True)

elif view=="Missing BOM Costs":
    st.dataframe(query_missing_costs(DEMO_CONN,dt_start,dt_end,boards_sel,mos_sel,machines_sel,components_sel,bl),use_container_width=True)

with st.expander("⬇️ Export to Excel"):
    if st.button("Prepare Excel Report", type="secondary", key="prepare_excel_report"):
        with st.spinner("Building Excel report..."):
            try:
                from openpyxl import Workbook
                from openpyxl.utils.dataframe import dataframe_to_rows
            except ImportError:
                st.error("openpyxl is required for Excel export. Add it to requirements.txt.")
                st.stop()
            wb = Workbook()
            # Sheet 1: Summary
            ws1 = wb.active
            ws1.title = "Summary"
            smd = query_summary(DEMO_CONN, dt_start, dt_end, boards_sel, mos_sel, machines_sel, components_sel, bl)
            if not smd.empty:
                smd_out = smd.drop(columns=["TotalPlacementCost", "SuccessfulCount"], errors="ignore").reset_index(drop=True)
                for r in dataframe_to_rows(smd_out, index=False, header=True):
                    ws1.append(r)
            # Sheet 2: Spit Events
            ws2 = wb.create_sheet("Spit Events")
            evf = query_events(DEMO_CONN, dt_start, dt_end, boards_sel, mos_sel, machines_sel, components_sel, bl)
            if not evf.empty:
                evf2 = evf[evf["RejectCode"].isin(REJECT_CODES)].reset_index(drop=True)
                for r in dataframe_to_rows(evf2, index=False, header=True):
                    ws2.append(r)
            # Sheet 3: Spit to Success % by Component
            ws3 = wb.create_sheet("Spit to Success %")
            rtf = query_success_ratio(DEMO_CONN, dt_start, dt_end, boards_sel, mos_sel, machines_sel, components_sel, bl)
            if not rtf.empty:
                rt_out = rtf[["Component","Description","SuccessfulCount","Spits","SpitToSuccessPct"]].sort_values(
                    ["SpitToSuccessPct","Spits"], ascending=[False,False]).reset_index(drop=True)
                for r in dataframe_to_rows(rt_out, index=False, header=True):
                    ws3.append(r)
            # Sheet 4: Successful Placements
            ws4 = wb.create_sheet("Successful Placements")
            spf = query_successful_placements(DEMO_CONN, dt_start, dt_end, boards_sel, mos_sel, machines_sel, components_sel, bl)
            if not spf.empty:
                spf["UnitCost"] = spf["Component"].map(lambda c: float(bl.get(c.strip().upper(), 0.0)))
                spf["TotalCost"] = spf["SuccessfulCount"] * spf["UnitCost"]
                tpc_val = query_total_placement_cost(DEMO_CONN, dt_start, dt_end, boards_sel, mos_sel, machines_sel, bl)
                spf["TotalPlacementCost"] = tpc_val
                cols = [c for c in spf.columns if c != "TotalPlacementCost"] + ["TotalPlacementCost"]
                spf = spf[cols]
                for r in dataframe_to_rows(spf, index=False, header=True):
                    ws4.append(r)
            from openpyxl.worksheet.table import Table, TableStyleInfo
            from openpyxl.utils import get_column_letter
            def _format_sheet(ws, start_row=1, end_row=None):
                """Auto-fit column widths and add table formatting from start_row to end_row."""
                max_r = end_row if end_row is not None else ws.max_row
                if max_r < start_row + 1 or ws.max_column < 1:
                    return
                for col_idx in range(1, ws.max_column + 1):
                    max_width = 0
                    for row_idx in range(start_row, max_r + 1):
                        cell_val = ws.cell(row=row_idx, column=col_idx).value
                        if cell_val is not None:
                            max_width = max(max_width, len(str(cell_val)))
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width + 4, 55)
                tbl_ref = f"A{start_row}:{get_column_letter(ws.max_column)}{max_r}"
                import random as _rand
                tbl = Table(displayName=f"{ws.title.replace(' ', '')}_R{start_row}_{_rand.randint(1000,9999)}", ref=tbl_ref)
                tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                ws.add_table(tbl)
            # Format all sheets
            for ws in [ws1, ws2, ws3, ws4]:
                _format_sheet(ws, 1)
            import io as _io
            bio = _io.BytesIO()
            wb.save(bio)
            st.session_state["export_bytes"] = bio.getvalue()
    if st.session_state.get("export_bytes"):
        fname = f"proline_demo_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        st.download_button("📥 Download Excel Report", data=st.session_state["export_bytes"], file_name=fname, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

st.markdown("---")
st.caption("📣 This is a demonstration version using synthetic data. Contact nabil@pylinesolutions.co.uk to get the full version with your factory's real data.")
